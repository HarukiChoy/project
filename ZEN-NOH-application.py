import tkinter as tk
from tkinter import *
import time
import datetime
from tkinter import simpledialog as sd
from tkinter import ttk
import pandas as pd
import numpy as np
from glob import glob
from pathlib import Path
import os


def MainProgram(date, name_input, password_input):
    def Login():
        DownList = chrome.find_element_by_id("dropdownlistContentsubcomps")
        DownList.click()
        time.sleep(0.5)
        HK = chrome.find_element_by_xpath(
            '//*[@id="listitem1innerListBoxsubcomps"]/span')
        HK.click()
        time.sleep(0.5)
        User = chrome.find_element_by_id("username")
        User.send_keys(name_input)
        time.sleep(0.5)
        Password = chrome.find_element_by_id("password")
        Password.send_keys(password_input)
        time.sleep(0.5)
        LoginButton = chrome.find_element_by_id("loginButton")
        LoginButton.click()
        time.sleep(0.5)
        alert = Alert(chrome)
        alert.accept()
        time.sleep(2)
    #________________________________________________________________________________________#

    def Invoice():
        id10 = chrome.find_element_by_id("10")
        id10.click()
        time.sleep(0.5)
        id1003 = chrome.find_element_by_id("1003")
        id1003.click()
        time.sleep(1)
    #________________________________________________________________________________________#

    def Summarize(df, date):
        R = df['備註'].str.split('\n', expand=True)
        for i in range(len(R)):
            if R.iloc[i, 2] == 'EG':
                df.loc[i, 'EgDelivery'] = True
            else:
                df.loc[i, 'EgDelivery'] = False
        df['COD'] = R.iloc[:, 1]
        df['送貨地點'] = R.iloc[:, 0]
        df = df.loc[df['EgDelivery'] == True]

        column = ['發票單號', '發票日期', '客戶名稱', '送貨地點', 'COD']
        _df = df[column].reset_index()

        y = df['貨品內容'].str.split('\n', expand=True)
        y = y.apply(lambda x: x.str.split('\t')).reset_index()

        for w in range(len(y)):
            for e in range(len(y.columns)):
                try:
                    y[e][w][2] = y[e][w][2].split('.')[0]
                except:
                    pass
        y = y.iloc[:, 1:]
        concat = pd.concat([_df, y], axis=1)
        concat = concat.drop(columns='index')

        # Summary
        NewPd = pd.DataFrame()
        for k in range(concat.shape[0]):
            for i in range(concat.shape[1]-5):
                if not concat.loc[k, i] == None:
                    NewPd = pd.concat([NewPd, pd.DataFrame(
                        [concat.loc[k, '發票單號'],
                         concat.loc[k, '發票日期'],
                         concat.loc[k, '客戶名稱'],
                         concat.loc[k, '送貨地點'],
                         concat.loc[k, 'COD'],
                         concat.loc[k, i]
                         ]).T], axis=0
                    )
        NewPd = NewPd.reset_index()
        NewPd = NewPd.drop(columns='index')

        for h in range(NewPd.shape[0]):
            NewPd.loc[h, 'Item Code'] = NewPd.iloc[h, 5][0]
            NewPd.loc[h, 'Item Name'] = NewPd.iloc[h, 5][1]
            NewPd.loc[h, 'Quantity'] = float(NewPd.iloc[h, 5][2])
        NewPd = NewPd.drop(columns=[5], axis=1)
        NewPd = NewPd.rename(
            columns={0: 'Inv#', 1: 'Date', 2: 'Customer', 3: 'Location', 4: 'COD'})
        paths = os.path.abspath(os.getcwd())
        path = f'{paths}/OUTPUT_{date}'
        os.mkdir(path)
        NewPd.to_excel(f'{path}/Summary_{date}.xlsx', index=False)

        # Delivery Allocation
        for k in range(concat.shape[0]):
            count = 0
            for i in range(concat.shape[1]-4):
                try:
                    count += int(concat.loc[k, i][2])
                except:
                    continue
            concat.loc[k, '送貨數量'] = count
        column2 = ['客戶名稱', '送貨地點', '送貨數量']
        _df2 = concat[column2]
        _df2 = _df2.groupby(['客戶名稱', '送貨地點']).sum()
        _df2.to_excel(f'{path}/Delivery_Allocation_{date}.xlsx', index=True)

        # EG_LM
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side, Font, Color, colors, PatternFill
        df_LD = pd.read_excel(f'{path}/Summary_{date}.xlsx')
        _df_LD = df_LD[['Customer', 'Location',
                        'Item Name', 'Quantity', 'Item Code']]
        _df_LD['Customer'] = _df_LD['Customer'].replace(
            "Pan Pacific Retail Management (HK) Co., Ltd", 'DONKI')
        _df_LD['Customer'] = _df_LD['Customer'].replace(
            "Dah Chong Hong Ltd", '大昌')
        _df_LD['Customer'] = _df_LD['Customer'].replace("Yata Ltd", 'YATA')
        _df_LD['Customer'] = _df_LD['Customer'].replace("彩鷗國際有限公司", '優品360')
        _df_LD['Customer'] = _df_LD['Customer'].replace("西原商會香港有限公司", '西原商會')
        _df_LD['Customer'] = _df_LD['Customer'].replace("日健日本食品有限公司", '日健')

        _df_LD2 = _df_LD.pivot_table(index=['Customer', 'Location'], columns=[
                                     'Item Name', 'Item Code'], aggfunc='sum').reset_index()
        _df_LD2.to_excel('PivotDraft.xlsx')
        filepath = glob('PivotDraft.xlsx')
        file = filepath[0]
        wb = load_workbook(file)
        ws = wb.active
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws['A4'] = "Lot. No."
        ws['A3'].font = Font(bold=True, size=13)
        ws['A4'].font = Font(bold=True, size=13)
        ws['A2'].font = Font(bold=True, size=13)
        for i in ws[1]:
            i.font = Font(bold=True, size=18)
        YellowFill = PatternFill(
            start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        for k in ws[4]:
            k.fill = YellowFill
        for row in ws.iter_rows(min_row=1, min_col=1):
            for cell in row:
                cell.border = border
        # Align Wrap Text
        for cell in ws[2]:
            cell.alignment = cell.alignment.copy(wrapText=True)
        # Adjust the column width of column B and C
        x = [len(i.value) for i in ws['B'] if i.value != None]
        y = [len(i.value) for i in ws['C'] if i.value != None]
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = max(x)+2
        ws.column_dimensions['C'].width = max(y)+2
        wb.save(f"{path}/Pivot_LM_{date}.xlsx")

        # EG_LD
        df3 = pd.read_excel(f"{path}/Pivot_LM_{date}.xlsx")
        df3 = df3.T
        df3.to_excel('Pivot_LD_Draft.xlsx', index=False)
        wb2 = load_workbook('Pivot_LD_Draft.xlsx')
        ws2 = wb2.active
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Add Border
        for row in ws2.iter_rows(min_row=1, min_col=1):
            for cell in row:
                cell.border = border
        # Align Wrap Text
        for cell in ws2[3]:
            cell.alignment = cell.alignment.copy(wrapText=True)
        for cell in ws2[4]:
            cell.alignment = cell.alignment.copy(wrapText=True)
        ws2.delete_rows(1)
        for k in ws2['C']:
            k.fill = YellowFill
        # Adjust the column width of column A and B
        x = [len(i.value) for i in ws2['A'] if i.value != None]
        y = [len(i.value) for i in ws2['B'] if i.value != None]
        ws2.column_dimensions['A'].width = max(x)+2
        ws2.column_dimensions['B'].width = max(y)+2
        wb2.save(f"{path}/Pivot_LD_{date}.xlsx")

        LDDraft = Path('Pivot_LD_Draft.xlsx')
        LDDraft.unlink()
        PivotDraft = Path('PivotDraft.xlsx')
        PivotDraft.unlink()

   #________________________________________________________________________________________#

    def SendingMail(date):
        Body = """
        Dear EG Team,

        Good day!
        Thank you always of your delivery arrangement.
        Enclosed please find the delivery summary for your reference.
        Please noted that there might be some of the delivery orders are not included in it.

        Should you have any enquiry, please do not hesitate to contact us at once.
        Thank you.

        Best Regards,
        ZEN-NOH INTERNATIONAL HONG KONG LIMITED
        """

        import win32com.client as win32
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = 'xxx@xxx.com' + ';' + 'yyy@xxx.com'
        mail.cc ='xxx@xxx.com' + ';' + 'yyy@xxx.com' 

        mail.Subject = f'Delivery Summary of {date}'
        mail.Body = f'{Body}'
        mail.Importance = 2

        ExcelPaths = glob('OUTPUT_{}/*.xlsx'.format(date))
        ExcelPath = ExcelPaths[0]
        ExcelPath_1 = ExcelPaths[1]
        ExcelPath_2 = ExcelPaths[2]
        ExcelPath_3 = ExcelPaths[3]
        paths = os.path.abspath(os.getcwd())
        Attachment = r'{}\{}'.format(paths, ExcelPath)
        Attachment_1 = r'{}\{}'.format(paths, ExcelPath_1)
        Attachment_2 = r'{}\{}'.format(paths, ExcelPath_2)
        Attachment_3 = r'{}\{}'.format(paths, ExcelPath_3)
        mail.Attachments.Add(Attachment)
        mail.Attachments.Add(Attachment_1)
        mail.Attachments.Add(Attachment_2)
        mail.Attachments.Add(Attachment_3)

        mail.Send()
    #________________________________________________________________________________________#
    # Main Program
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.alert import Alert
    from webdriver_manager.chrome import ChromeDriverManager
    import pandas as pd
    import numpy as np
    from glob import glob
    from pathlib import Path
    import os

    options = Options()
    options.add_argument("--disable-notifications")
    chrome = webdriver.Chrome(
        ChromeDriverManager().install(), chrome_options=options)
    url = 'http://103.20.63.146:8080/ciderp2019/cust/zennoh/'
    chrome.implicitly_wait(5)
    chrome.get(url)
    Login()  # Login into Cider System
    Invoice()  # Click into the Invoice Page

    # Change to the activated window
    windows = chrome.window_handles
    chrome.switch_to.window(windows[-1])
    time.sleep(2)
    ChangeDate = chrome.find_element_by_xpath(
        "//html/body/div[3]/div[1]/form/table/tbody/tr[1]/td[4]/div/div/div[1]/div")
    ChangeDate.click()
    time.sleep(1)
    LastDate = chrome.find_element_by_xpath(
        "//html/body/div[53]/div[1]/div/table/tbody/tr[2]/td[2]/table/tbody/tr[6]/td[7]")
    LastDate.click()
    time.sleep(1)
    GO = chrome.find_element_by_xpath(
        "//html/body/div[3]/div[1]/form/table/tbody/tr[1]/td[5]/input")
    GO.click()
    time.sleep(1)
    InvDate = chrome.find_element_by_xpath(
        "//html/body/div[6]/div/div/div/div/div[4]/div[1]/div[2]/div/div[19]/input")
    InvDate.send_keys(date)
    time.sleep(1)
    ExportExcel = chrome.find_element_by_id("btnExcel")
    ExportExcel.click()
    time.sleep(1)
    chrome.quit()

    # Go to the Excel file part
    home = str(Path.home())
    Downloads = f'{home}\\Downloads'

    filepaths = glob(f'{Downloads}\*.xls')
    filepath = filepaths[0]
    df = pd.read_excel(filepath)

    Summarize = Summarize(df, date)
    Xls1 = Path(f'{filepath}')
    Xls1.unlink()
    time.sleep(1)

    SendingMail(date)
#________________________________________________________________________________________#


def Inventory():
    def GetEgFile():
        User = chrome.find_element_by_id("usrname")
        User.send_keys("xxx")
        time.sleep(0.5)
        Password = chrome.find_element_by_id("usrpass")
        Password.send_keys("yyy")
        time.sleep(0.5)
        LoginButton = chrome.find_element_by_id("btn_login")
        LoginButton.click()
        time.sleep(10)
        url = 'http://www.e-gain.com.hk:8088/ittms/ittms.php/ctrl_lot_enquiry'
        chrome.get(url)
        time.sleep(10)
        Search = chrome.find_element_by_id("f_lot_enquiry_btn_search")
        Search.click()
        time.sleep(10)
        chrome.execute_script("window.scrollTo(0,document.body.scrollHeight)")
        time.sleep(5)
        Export = chrome.find_element_by_xpath(
            '//*[@id="table_lot_enquiry_browse_pager"]/div/span[1]/span/span[2]')
        Export.click()
        time.sleep(5)
        chrome.quit()

    def TurnXlsToXlsx():
        import xlwings as xw
        home = str(Path.home())
        Downloads = f'{home}\\Downloads'
        filepaths = glob(f'{Downloads}\*.xls')
        filepath = filepaths[0]
        app = xw.App(visible=True, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        wb = app.books.open(filepath)
        wb.save('EG.xlsx')
        wb.close()
        app.quit()
        Xls2 = Path(f'{filepath}')
        Xls2.unlink()
        time.sleep(1)

    def SortingFile():
        ExcelPaths = glob('EG.xlsx')
        ExcelPath = ExcelPaths[0]
        df = pd.read_excel(ExcelPath)
        df = df.iloc[3:, :].T.reset_index()
        df = df.iloc[:, 1:].T
        UseCol = [0, 1, 3, 23, 6]
        _df = df[UseCol]
        _df = _df.rename(columns={0: 'Item No.', 1: 'Receive Date', 3: 'Lot No.',
                         23: 'Description', 6: 'Available Quantity'})
        _df2 = df[12]
        _df2 = _df2.str.split(' ', expand=True)
        _df2 = _df2.rename(columns={0: 'Expiry Date'})
        _df2 = _df2['Expiry Date']
        df = pd.concat([_df, _df2], axis=1)
        df = df[['Item No.', 'Receive Date', 'Lot No.',
                 'Description', 'Expiry Date', 'Available Quantity']]
        df = df.set_index(['Receive Date', 'Lot No.'])
        df = df.sort_values(['Receive Date', 'Lot No.'])
        df.to_excel('Draft.xlsx', index=True)
        EG = Path('EG.xlsx')
        EG.unlink()

    def AddBorder():
        import datetime
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side, Font, Color, colors, PatternFill
        today = datetime.date.today()
        filepath = glob('Draft.xlsx')
        file = filepath[0]
        wb = load_workbook(file)
        ws = wb.active
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_col=6):
            for cell in row:
                cell.border = border
        wb.save(f"Inventory_{today}.xlsx")
        draft = Path('Draft.xlsx')
        draft.unlink()

    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.alert import Alert
    from webdriver_manager.chrome import ChromeDriverManager
    options = Options()
    options.add_argument("--disable-notifications")
    chrome = webdriver.Chrome(
        ChromeDriverManager().install(), chrome_options=options)
    url = 'http://www.e-gain.com.hk:8088/ittms/ittms.php'
    chrome.implicitly_wait(5)
    chrome.get(url)
    time.sleep(1)

    GetEgFile()
    time.sleep(2)

    TurnXlsToXlsx()
    time.sleep(2)

    SortingFile()
    time.sleep(2)

    AddBorder()
#________________________________________________________________________________________#


def Kamigumi():
    def docReader(path):
        word = win32.Dispatch("Word.Application")
        word.Visible = False
        wd = word.Documents.Open(path)
        doc = word.ActiveDocument
        paras = doc.Range().text
        return paras

    def choose_directory():
        directory_root = tk.Tk()
        directory_root.withdraw()
        path_work = askdirectory()
        return path_work

    from tkinter.filedialog import askdirectory

    choose = choose_directory()
    os.chdir(choose)
    filepaths = glob("*.doc")
    _df = pd.DataFrame()
    Final = pd.DataFrame()
    for i in range(len(filepaths)):
        filepath = filepaths[i]
        if filepath.startswith('d') or filepath.startswith('D'):
            paths = os.path.abspath(os.getcwd())
            path = f'{paths}\{filepath}'
            doc_out = docReader(path)
            list1 = doc_out.split('\r')
            subtotal = list1.index('SUBTOTAL')-1
            FirstOne = list1.index('AMOUNT')+1
            list2 = list1[FirstOne:subtotal]
            list3 = [index for index, element in enumerate(
                list2) if element == '']
            _df = pd.DataFrame()
            for i in list3:
                x = np.array(list2[i+1:i+8])
                x = pd.DataFrame(x).T
                _df = pd.concat([_df, x])
                _df['InvNo'] = list1[14]
                _df['Vessel'] = list1[36]
                _df['Voy No'] = list1[35]
                _df['MASTER B/L NO.'] = list1[33]
                _df['Loading Port'] = list1[47]
                _df['ETD'] = list1[44]
                _df['ETA'] = list1[45]
                _df['Volume'] = list1[59]
                _df['Package'] = list1[60]
                # _df['Weight'] = list1[58].replace(',', '')
                _df['Weight'] = float(list1[58].replace(',', ''))
                # _df['CBM'] = list1[56].replace(',', '')
                _df['CBM'] = float(list1[56].replace(',', ''))
                _df['Container No'] = list1[71]
            _df[1] = _df[1].astype(float)
            _df[3] = _df[3].astype(float)
            _df[4] = (_df[4].str.replace(',', '')).astype(float)
            _df[5] = _df[5].str.split(' ', expand=True)[1]
            _df[5] = (_df[5].str.replace(',', '')).astype(float)
            cols = ['InvNo', 'Vessel', 'Voy No', 'Loading Port', 'ETD', 'ETA', 'MASTER B/L NO.', 'Volume', 'Package',
                    'Weight', 'CBM', 'Container No', 'ITEM', 'UNIT', 'QTY', 'CURR.', 'EX.RATE', 'UNIT PRICE', 'AMOUNT']
            _df = _df.rename(columns={0: 'UNIT', 1: 'QTY', 2: 'CURR.',
                             3: 'EX.RATE', 4: 'UNIT PRICE', 5: 'AMOUNT', 6: 'ITEM'})
            df = _df[cols].reset_index()
            df = df.iloc[:, 1:]
        Final = pd.concat([Final, df], axis=0)
    Final.to_excel('Kamigumi_Summary.xlsx', index=False)
#________________________________________________________________________________________#


def GetInput():
    name_input = UserEntry.get()
    password_input = PassEntry.get()
    date = combo_1.get()
    if name_input == '':
        tk.messagebox.showwarning(
            title="Username Warning", message="You haven't filled in the Username")
    if password_input == '':
        tk.messagebox.showwarning(
            title="Password Warning", message="You haven't filled in the Password")
    if name_input and password_input != '':
        return MainProgram(date, name_input, password_input)
#________________________________________________________________________________________#


def get_time():
    global time1
    time1 = ''
    time2 = time.strftime('%Y-%m-%d %H:%M:%S')
    tk.Label(root, text='The time now is :', font=("Times", 14),
             fg='Red').grid(row=10, column=4, sticky=tk.W)
    if time2 != time1:
        time1 = time2
        clock = tk.Label(root, text=time1, font=("Times", 15), fg='Red')
        clock.configure(text=time2)
        clock.grid(row=10, column=5, sticky=tk.E)
        clock.after(200, get_time)
#________________________________________________________________________________________#


def EggStock():
    import pygsheets

    def clear_treeview():
        tree.delete(*tree.get_children())
    gc = pygsheets.authorize(
        service_file='./probable-analog-345616-96dd881a0da3.json')
    sht = gc.open_by_url(
        'https://docs.google.com/spreadsheets/d/1JZgokI9lg5FQZxBmzKhL2diC7het0DJDIfzFPLtXybw/edit#gid=0')
    wks = sht[0]
    df = pd.DataFrame(wks.get_all_records())

    for i in tree.get_children():
        tree.delete(i)
    root.update()
    tree["column"] = list(df.columns)
    tree["show"] = "headings"
    for col in tree["column"]:
        tree.heading(col, text=col)
    df_rows = df.to_numpy().tolist()
    for row in df_rows:
        tree.insert("", "end", values=row)
    tree.grid()
#________________________________________________________________________________________#


root = tk.Tk()
root.title("ZEN-NOH TOOL BOX")
root.resizable(0, 0)

Menubar = tk.Menu(root)
root.config(menu=Menubar)
FileMenu = tk.Menu(Menubar, tearoff=False)
FileMenu.add_command(label="Exit", command=lambda: root.destroy())
Menubar.add_cascade(label="Exit", menu=FileMenu)

LabelFrame_0 = tk.LabelFrame(
    root, text="CIDER Login Info", font=("Times", 14), fg="Purple")
LabelFrame_0.grid(row=0, column=0, columnspan=3, rowspan=2,
                  sticky=tk.E+tk.W, padx=10, pady=10)
UserName = tk.Label(LabelFrame_0, text="Username: ", fg="Blue")
UserName.grid(row=0, column=0, sticky=tk.W, pady=1)
UserEntry = tk.Entry(LabelFrame_0, width=20, fg="Purple")
UserEntry.grid(row=0, column=1, sticky=tk.E, pady=1)

Password = tk.Label(LabelFrame_0, text="Password: ", fg="Blue")
Password.grid(row=1, column=0, sticky=tk.W, pady=1)
PassEntry = tk.Entry(LabelFrame_0, width=20, show="*", fg="Purple")
PassEntry.grid(row=1, column=1, sticky=tk.E, pady=1)

label_1 = tk.Label(LabelFrame_0, text="Choose invoice date: ", fg="Blue")
label_1.grid(row=2, column=0, sticky=tk.W, pady=1)
ListOfDate = []
today = datetime.date.today()
for i in range(7):
    date_1 = today + datetime.timedelta(days=i)
    ListOfDate.append(date_1)
combo_1 = ttk.Combobox(LabelFrame_0, values=ListOfDate)
combo_1.current(1)
combo_1.grid(row=2, column=1, sticky=tk.E, pady=1)


LabelFrame_1 = tk.LabelFrame(
    root, text="EVERGAIN Delivery Summary", font=("Times", 14), fg="Purple")
LabelFrame_1.grid(row=0, column=3, columnspan=3, rowspan=2, padx=10, pady=10)
btn_2 = tk.Button(LabelFrame_1, text='Download and Send Summary',
                  command=lambda: GetInput(), font=("Times", 12), bg="Pink", fg="Blue", width=27)
btn_2.grid(row=1, column=0, pady=2)


LabelFrame_2 = tk.LabelFrame(
    root, text="EVERGAIN Inventory", font=("Times", 14), fg="Purple")
LabelFrame_2.grid(row=2, column=0, columnspan=2, padx=10, pady=10)
btn_3 = tk.Button(LabelFrame_2, text='GET EG Inventory', command=lambda: Inventory(
), font=("Times", 12), bg="Pink", fg="Blue", width=27)
btn_3.grid(row=0, column=0, pady=2)

LabelFrame_4 = tk.LabelFrame(
    root, text="Google Spread Sheet Linking", font=("Times", 14), fg="Purple")
LabelFrame_4.grid(row=2, column=2, columnspan=2, padx=10, pady=10)
btn_3 = tk.Button(LabelFrame_4, text='SHOW Stock', command=lambda: EggStock(
), font=("Times", 12), bg="Pink", fg="Blue", width=27)
btn_3.grid(row=0, column=0, pady=2)

LabelFrame_3 = tk.LabelFrame(
    root, text="KAMIGUMI Invoice Summary", font=("Times", 14), fg="Purple")
LabelFrame_3.grid(row=2, column=4, columnspan=2, padx=10, pady=10)
btn_4 = tk.Button(LabelFrame_3, text='Export Summary', command=lambda: Kamigumi(
), font=("Times", 12), bg="Pink", fg="Blue", width=27)
btn_4.grid(row=0, column=0, pady=2)

Frame_4 = tk.Frame(root, width=400, height=200, relief=tk.RIDGE, borderwidth=2)
Frame_4.grid(row=3, columnspan=6, sticky=tk.E+tk.W, padx=10, pady=10)
tree = ttk.Treeview(Frame_4)

get_time()
root.mainloop()
